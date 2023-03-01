#responsible: agata.mastalska@globallogic.com
#location: Wroclaw

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import unicorn
from core.basetest import BaseTest
import re
from dstl.auxiliary.init import dstl_detect
# from dstl.auxiliary.restart_module import dstl_restart


class Test(BaseTest):
    """
    Test checks for various parameters and generates excel file
    1. Prepare excel file 
    2. read params about firmware version, bootloader version, revision, radioband etc
    3. store those params on excel file in logs folder 
    """

    def setup(test):
        test.dut.dstl_detect()

    def run(test):
        wb = openpyxl.Workbook()
        ws = wb.active

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        for col in range(5):
            for row in range(12):
                ws.cell(column=(col + 1), row=(row + 1)).border = thin_border
                ws.cell(column=(col + 1), row=(row + 1)).alignment = Alignment(vertical="center", wrap_text=True)
                ws.cell(column=(col + 1), row=(row + 1)).font = Font(size=9, name="Arial")
                if col == 1:
                    ws.cell(column=2, row=(row + 1)).value = "SW"
                    ws.cell(column=2, row=(row + 1)).font = Font(bold=True)
                if col == 4:
                    ws.cell(column=5, row=(row + 1)).value = "name_param"

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30

        ws["A1"] = "Internal software Version nummer (build\n\tID SW Version, AT^S…)"
        test.expect(test.dut.at1.send_and_verify("at^sos=ver", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"ver[\r\n]+([ \r\n\-\w.\",\:@]+)[\r\n]+OK")
        ws['C1'] = result[0]
        ws["D1"] = result[1]
        ws["E1"] = "at^sos=ver"

        ws["A2"] = "Bootloader Version"
        test.expect(test.dut.at1.send_and_verify("at^sos=bootloader/info", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"[bB]ootloader[_ ]version[ _]name ?: (\d+.\d+(_[A-Z0-9]+)?)")
        ws["C2"] = result[0]
        ws["D2"] = result[1]
        ws["E2"] = "at^sos=bootloader/info"

        ws["A3"] = "Software Version Number\n(SVN to SW Version)"
        test.expect(test.dut.at1.send_and_verify("at^cicret=swn", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"\n([\w-]+) RELEASE,")
        ws["C3"] = result[0]
        ws["D3"] = result[1]
        ws["E3"] = "at^cicret=swn"

        ws["A4"] = "Software Revision Number\n(Rev. to SW Version, ATI)"
        ws["A6"] = "Application Revision String (ATI1)"
        test.expect(test.dut.at1.send_and_verify("ati1", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(REVISION \d{2}.\d{3})", r"(A-REVISION [0-9.]+)")
        ws["C4"] = result[0]
        ws["E4"] = "ati1"
        ws["C6"] = result[1]
        ws["D4"] = result[2]
        ws["D6"] = result[2]
        ws["E6"] = "ati1"

        ws["A5"] = "Product identification (ATI)"
        test.expect(test.dut.at1.send_and_verify("ati", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(REVISION \d{2}.\d{3})")
        ws["C5"] = result[0]
        ws["D5"] = result[1]
        ws["E5"] = "ATI"

        ws["A7"] = "Engineering Change Number (ATI255)"
        test.expect(test.dut.at1.send_and_verify("ati255", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"\r\n(\w+)\r\n")
        ws["C7"] = result[0]
        ws["D7"] = result[1]
        ws["E7"] = "ati255"

        ws["A8"] = "Check Standard Mapping (ATI281)"
        test.expect(test.dut.at1.send_and_verify("ati281", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"\r\n([\w\-]+)\r\n")
        ws["C8"] = result[0]
        ws["D8"] = result[1]
        ws["E8"] = "ati281"

        ws["A9"] = "Radio Band Setting (scfg=Radio/Band)"
        test.expect(test.dut.at1.send_and_verify('at^scfg="Radio/band"', '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(\^SCFG: \"Radio/Band\",\"\d+\")")
        if result[1] != "OK":
            test.expect(test.dut.at1.send_and_verify('at^scfg="Radio/Band/3G"', '.*OK.*'))
            response = test.dut.at1.last_response
            result_3g = get_result_of_at_command(test, response, r"(\^SCFG: \"Radio/Band/3G\",\"\w+\")")

            test.expect(test.dut.at1.send_and_verify('at^scfg="Radio/Band/4G"', '.*OK.*'))
            response = test.dut.at1.last_response
            result_4g = get_result_of_at_command(test, response, r"(\^SCFG: \"Radio/Band/4G\",\"\w+\")")

            if result_3g[1] != "OK" and result_4g[1] == "OK":
                ws["C9"] = result_4g[0]
                ws["D9"] = "OK"
            elif result_3g[1] == "OK" and result_4g[1] != "OK":
                ws["C9"] = result_3g[0]
                ws["D9"] = "OK"
            elif result_3g[1] == "OK" and result_4g[1] == "OK":
                ws["C9"] = "{}\r\n{}".format(result_3g[0], result_4g[0])
                ws["D9"] = "OK"
            else:
                ws["D9"] = result_3g[1]
        else:
            ws["C9"] = result[0]
            ws["D9"] = result[1]
            ws["E9"] = 'at^scfg="Radio/band"'

        ws["A10"] = "Variant check (e.g. SNBTA, log into\nnetwork with mit CI-Module, SVN (Luft))"
        test.expect(test.dut.at1.send_and_verify("ati176", '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(\d{15}.(\d{2}))")
        ws["C10"] = result[0]
        ws["D10"] = result[1]
        ws["E10"] = "ati176"

        ws["A12"] = "IMEI (AT+GSN) / serial number\n(AT^SGSN) read and check with label"
        test.expect(test.dut.at1.send_and_verify('at+gsn', '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(\d+)")
        ws["C12"] = result[0]
        ws["D12"] = result[-1]
        ws["E12"] = 'at+gsn'

        ws["A11"] = "Locks (CLCK)"
        test.expect(test.dut.at1.send_and_verify("at^sjam=4", '.*OK.*'))
        test.expect(test.dut.at1.send_and_verify("at^sjam=5", '.*OK.*'))
        response = test.dut.at1.last_response
        if response:
            test.expect(test.dut.at1.send_and_verify('at^sfsa="ls","a:/"', '.*OK.*'))
        test.expect(test.dut.at1.send_and_verify("ati8", '.*OK.*'))
        test.expect(test.dut.at1.send_and_verify("at^spow?", '.*OK.*'))
        test.expect(test.dut.at1.send_and_verify('AT+CPIN="9999"', '.*OK.*'))

        test.expect(test.dut.at1.send_and_verify('at+clck="SC",2', '.*OK.*'))
        response = test.dut.at1.last_response
        result = get_result_of_at_command(test, response, r"(CLCK: \d+)")
        ws["C11"] = result[0]
        ws["D11"] = result[-1]
        ws["E11"] = 'at+clck="SC",2'

        wb.save(f'{test.log_path}/reference_sample_check_{test.dut.software}.xlsx')
        test.log.info(f"excel saved: {test.log_path}/reference_sample_check_{test.dut.software}.xlsx")

    def cleanup(test):
        pass

if "__main__" == __name__:
    unicorn.main()

def get_result_of_at_command(test, response, *regex_patterns):
    to_return = []
    for regex_pattern in regex_patterns:
        result = re.search(regex_pattern, response)
        if result:
            to_return.append(result.group(1))
        else:
            to_return.append("")
    status = re.search(r"OK", response)
    if status:
        to_return.append("OK")
        return to_return
    result = re.search(r'CME ERROR: ([\w ]+)\r\n', response)
    if result:
        to_return.append("ERROR: {}".format(result.group(1)))
    else:
        to_return.append("ERROR")
    return to_return